"""
daily_report.py
---------------
Scans previous day's inbox (ALL emails, no filter), classifies them,
extracts structured fields from customer_requirement and quotation_rate_card
emails, builds a 2-sheet Excel, and emails it to REPORT_RECIPIENT.

Usage:
    python daily_report.py                        # yesterday's emails → email report
    python daily_report.py --dry-run              # save Excel locally, do not email
    python daily_report.py --since 2026/05/01     # backfill from date (uses subject filter + 1000 cap)

Schedule via Windows Task Scheduler at 8am daily (no args needed).

Required env vars:
    EMAIL_ACCOUNT, EMAIL_PASSWORD      (SMTP sender — labsxelta Gmail app password)
    REPORT_RECIPIENT                   (email address to receive daily report)
    GMAIL_MAILBOX                      (client inbox to scan)
    EMAIL_PROVIDER=gmail_workspace
"""

import json
import logging
import os
import re
import sys
from datetime import date
from io import BytesIO

from dotenv import load_dotenv
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

openai_client = OpenAI()

REPORT_RECIPIENT = os.environ.get("REPORT_RECIPIENT", "")
VALID_LABELS = {"customer_requirement", "quotation_rate_card", "general", "skip"}


# ---------------------------------------------------------------------------
# Email fetching
# ---------------------------------------------------------------------------

_LOGISTICS_SUBJECT_FILTER = (
    "subject:RFQ OR subject:quotation OR subject:\"rate request\" OR "
    "subject:\"rate card\" OR subject:\"spot rate\" OR subject:\"rate offer\" OR "
    "subject:\"rate sheet\" OR subject:\"rate update\" OR subject:\"freight rate\" OR "
    "subject:\"ocean freight\" OR subject:\"air freight\" OR subject:\"sea freight\" OR "
    "subject:enquiry OR subject:inquiry OR subject:\"quote request\" OR "
    "subject:\"booking request\" OR subject:\"rate req\" OR subject:FCL OR "
    "subject:LCL OR subject:\"20ft\" OR subject:\"40ft\" OR subject:\"20GP\" OR "
    "subject:\"40HC\" OR subject:cargo OR subject:shipment OR subject:\"rate valid\" OR "
    "subject:\"best rate\" OR subject:costing OR subject:tariff"
)


def _fetch_emails(since: str, before: str | None = None, backfill: bool = False) -> list[dict]:
    """Fetch emails between since and before (YYYY/MM/DD).

    Daily mode (backfill=False): fetches ALL emails, no subject filter, no cap.
    Backfill mode (backfill=True): applies subject filter + 1000 cap to handle large ranges.
    """
    provider = os.getenv("EMAIL_PROVIDER", "gmail").lower()

    if provider == "gmail_workspace":
        import time
        from gmail_connector import _get_service, _decode_header_value, _extract_body
        from googleapiclient.errors import HttpError

        query = f"after:{since}"
        if before:
            query += f" before:{before}"
        if backfill:
            query += f" ({_LOGISTICS_SUBJECT_FILTER})"

        cap = 1000 if backfill else 10000

        service = _get_service()
        refs = []
        page_token = None
        while len(refs) < cap:
            kwargs = {
                "userId": "me",
                "maxResults": min(500, cap - len(refs)),
                "q": query,
            }
            if page_token:
                kwargs["pageToken"] = page_token
            result = service.users().messages().list(**kwargs).execute()
            refs.extend(result.get("messages", []))
            page_token = result.get("nextPageToken")
            if not page_token:
                break

        logger.info("Found %d emails (%s)", len(refs), "backfill" if backfill else "daily")
        emails = []
        for i, ref in enumerate(refs, 1):
            for attempt in range(3):
                try:
                    msg = service.users().messages().get(
                        userId="me", id=ref["id"], format="full"
                    ).execute()
                    headers = {h["name"]: h["value"]
                               for h in msg.get("payload", {}).get("headers", [])}
                    body = _extract_body(msg.get("payload", {})).strip()
                    if body and len(body) >= 20:
                        emails.append({
                            "id": msg["id"],
                            "from": headers.get("From", ""),
                            "subject": _decode_header_value(headers.get("Subject", "")),
                            "body": body,
                            "date": headers.get("Date", ""),
                        })
                    break
                except HttpError as e:
                    logger.warning("Skipping %s: %s", ref["id"], e)
                    break
                except Exception as e:
                    if attempt < 2:
                        logger.warning("Retry %d for %s: %s", attempt + 1, ref["id"], e)
                        time.sleep(2 ** attempt)
                    else:
                        logger.error("Failed %s after 3 attempts: %s", ref["id"], e)
            if i % 100 == 0:
                logger.info("  fetched %d/%d", i, len(refs))
        return emails

    # Fallback: IMAP/Outlook
    from email_connector import fetch_latest_emails
    result = fetch_latest_emails(limit=200, offset=0)
    return result.get("emails", [])


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def _classify_email(email: dict) -> tuple[str, float]:
    """Classify one email. Returns (label, confidence)."""
    prompt = (
        "You are classifying an email from a freight/logistics inbox.\n\n"
        f"From: {email['from']}\n"
        f"Subject: {email['subject']}\n"
        f"Body (first 600 chars):\n{email['body'][:600]}\n\n"
        "Classify into exactly one label:\n"
        "- customer_requirement: customer or agent asking for shipping rates, booking, or freight quote\n"
        "- quotation_rate_card: carrier or agent providing freight rates or price breakdown\n"
        "- general: shipment status, tracking, documents, invoices, operational\n"
        "- skip: spam, unsubscribe, out-of-office, or not logistics-related\n\n"
        'Reply ONLY with JSON: {"label": "...", "confidence": 0.0}'
    )
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=60,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        data = json.loads(raw)
        label = data.get("label", "skip")
        if label not in VALID_LABELS:
            label = "skip"
        return label, float(data.get("confidence", 0.0))
    except Exception as e:
        logger.warning("Classify failed for '%s': %s", email.get("subject", "")[:40], e)
        return "skip", 0.0


# ---------------------------------------------------------------------------
# Field extraction
# ---------------------------------------------------------------------------

_CRQ_PROMPT = """\
Extract structured fields from this logistics enquiry email.
Return ONLY valid JSON with these exact keys (use null if not found):
{{
  "sender_name": "...",
  "company": "...",
  "pol": "...",
  "pod": "...",
  "container_type": "...",
  "container_count": "...",
  "cargo_type": "...",
  "incoterms": "...",
  "remarks": "..."
}}

From: {from_}
Subject: {subject}
Body:
{body}
"""

_RCQ_PROMPT = """\
Extract structured fields from this freight rate card / rate offer email.
Return ONLY valid JSON with these exact keys (use null if not found):
{{
  "carrier_agent": "...",
  "pol": "...",
  "pod": "...",
  "container_type": "...",
  "rate_usd": "...",
  "validity": "...",
  "transit_time": "...",
  "remarks": "..."
}}

From: {from_}
Subject: {subject}
Body:
{body}
"""


def _extract_fields(email: dict, label: str) -> dict:
    """Extract structured fields from a classified email."""
    prompt_tpl = _CRQ_PROMPT if label == "customer_requirement" else _RCQ_PROMPT
    prompt = prompt_tpl.format(
        from_=email["from"],
        subject=email["subject"],
        body=email["body"][:1200],
    )
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=300,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        fields = json.loads(raw)
        fields["subject"] = email["subject"]
        fields["from"] = email["from"]
        fields["date"] = email.get("date", "")
        return fields
    except Exception as e:
        logger.warning("Extract failed for '%s': %s", email.get("subject", "")[:40], e)
        return {
            "subject": email["subject"],
            "from": email["from"],
            "date": email.get("date", ""),
            "error": str(e),
        }


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="D9E1F2")


def _style_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 22


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def _add_rows(ws, rows: list[dict], headers: list[str]) -> None:
    for r_idx, row in enumerate(rows, 2):
        fill = _ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(headers, 1):
            val = row.get(key.lower().replace(" ", "_"), row.get(key, ""))
            cell = ws.cell(row=r_idx, column=c_idx, value=val or "")
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)


_CRQ_HEADERS = [
    "Date", "From", "Subject",
    "Sender Name", "Company",
    "POL", "POD",
    "Container Type", "Container Count",
    "Cargo Type", "Incoterms", "Remarks",
]

_RCQ_HEADERS = [
    "Date", "From", "Subject",
    "Carrier / Agent",
    "POL", "POD",
    "Container Type", "Rate (USD)", "Validity", "Transit Time", "Remarks",
]

_CRQ_KEYS = [
    "date", "from", "subject",
    "sender_name", "company",
    "pol", "pod",
    "container_type", "container_count",
    "cargo_type", "incoterms", "remarks",
]

_RCQ_KEYS = [
    "date", "from", "subject",
    "carrier_agent",
    "pol", "pod",
    "container_type", "rate_usd", "validity", "transit_time", "remarks",
]


def build_excel(crq_rows: list[dict], rcq_rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()

    # Sheet 1: Quote Requests
    ws1 = wb.active
    ws1.title = "Quote Requests"
    ws1.freeze_panes = "A2"
    _style_header(ws1, _CRQ_HEADERS)
    for r_idx, row in enumerate(crq_rows, 2):
        alt = PatternFill("solid", fgColor="D9E1F2") if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(_CRQ_KEYS, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=row.get(key, "") or "")
            if alt:
                cell.fill = alt
            cell.alignment = Alignment(wrap_text=True)
    _auto_width(ws1)

    # Sheet 2: Rate Cards Received
    ws2 = wb.create_sheet("Rate Cards Received")
    ws2.freeze_panes = "A2"
    _style_header(ws2, _RCQ_HEADERS)
    for r_idx, row in enumerate(rcq_rows, 2):
        alt = PatternFill("solid", fgColor="D9E1F2") if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(_RCQ_KEYS, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=row.get(key, "") or "")
            if alt:
                cell.fill = alt
            cell.alignment = Alignment(wrap_text=True)
    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Email sender (with attachment)
# ---------------------------------------------------------------------------

def _send_report(excel_bytes: bytes, today: str, crq_count: int, rcq_count: int) -> None:
    import smtplib
    import ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    account = os.environ["EMAIL_ACCOUNT"]
    password = os.environ["EMAIL_PASSWORD"]
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    recipient = os.environ["REPORT_RECIPIENT"]

    msg = MIMEMultipart()
    msg["From"] = account
    msg["To"] = recipient
    msg["Subject"] = f"Daily Logistics Report — {today}"

    body = (
        f"Hi,\n\n"
        f"Please find today's logistics inbox summary attached.\n\n"
        f"  • Quote Requests (from customers/agents): {crq_count}\n"
        f"  • Rate Cards Received (from carriers/agents): {rcq_count}\n\n"
        f"Report date: {today}\n\n"
        f"Regards,\nBhatia Shipping Copilot"
    )
    msg.attach(MIMEText(body, "plain"))

    attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    attachment.set_payload(excel_bytes)
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition",
        f'attachment; filename="logistics_report_{today}.xlsx"',
    )
    msg.attach(attachment)

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
        server.ehlo()
        server.starttls(context=context)
        server.login(account, password)
        server.sendmail(account, recipient, msg.as_string())
    logger.info("Report emailed to %s", recipient)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _parse_args() -> tuple[str, str | None, bool]:
    """Parse CLI args. Returns (since, before, backfill).

    Default (no --since): since=yesterday, before=today, backfill=False.
    With --since DATE:    since=DATE, before=None, backfill=True (large range).
    """
    for i, arg in enumerate(sys.argv):
        if arg == "--since" and i + 1 < len(sys.argv):
            return sys.argv[i + 1], None, True
    yesterday = (date.today() - date.resolution * 1).strftime("%Y/%m/%d")
    today_str = date.today().strftime("%Y/%m/%d")
    return yesterday, today_str, False


def main() -> None:
    dry_run = "--dry-run" in sys.argv
    since, before, backfill = _parse_args()

    # Report label = yesterday's date for daily runs, today for backfill
    report_date = since.replace("/", "-") if backfill else (date.today() - date.resolution * 1).strftime("%Y-%m-%d")

    logger.info("=== Daily Report — %s ===", report_date)
    if backfill:
        logger.info("Backfill mode: since %s (subject filter + 1000 cap active)", since)
    else:
        logger.info("Daily mode: %s only, ALL emails, no filter", since)

    emails = _fetch_emails(since, before=before, backfill=backfill)
    if not emails:
        logger.info("No emails found. Nothing to report.")
        return

    logger.info("Classifying %d emails...", len(emails))
    crq_emails, rcq_emails = [], []
    for i, em in enumerate(emails, 1):
        label, conf = _classify_email(em)
        logger.info("[%3d/%d] %-26s conf=%.2f  %s", i, len(emails), label, conf, em["subject"][:55])
        if label == "customer_requirement" and conf >= 0.75:
            crq_emails.append(em)
        elif label == "quotation_rate_card" and conf >= 0.75:
            rcq_emails.append(em)

    logger.info("Extracting fields from %d quote requests...", len(crq_emails))
    crq_rows = [_extract_fields(em, "customer_requirement") for em in crq_emails]

    logger.info("Extracting fields from %d rate cards...", len(rcq_emails))
    rcq_rows = [_extract_fields(em, "quotation_rate_card") for em in rcq_emails]

    logger.info("Building Excel...")
    excel_bytes = build_excel(crq_rows, rcq_rows)

    filename = f"logistics_report_{report_date}.xlsx"

    if dry_run:
        with open(filename, "wb") as f:
            f.write(excel_bytes)
        logger.info("DRY RUN — saved to %s (not emailed)", filename)
        logger.info("  Quote Requests : %d", len(crq_rows))
        logger.info("  Rate Cards     : %d", len(rcq_rows))
        return

    if not REPORT_RECIPIENT:
        logger.error("REPORT_RECIPIENT not set in .env — cannot email report.")
        with open(filename, "wb") as f:
            f.write(excel_bytes)
        logger.info("Saved locally to %s", filename)
        return

    _send_report(excel_bytes, report_date, len(crq_rows), len(rcq_rows))
    logger.info("Done. Quote Requests: %d | Rate Cards: %d", len(crq_rows), len(rcq_rows))


if __name__ == "__main__":
    main()

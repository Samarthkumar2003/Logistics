"""
daily_report.py
---------------
Summarises a day's freight enquiries and rate cards into a 2-sheet Excel
workbook and emails it to REPORT_RECIPIENT.

Reads from the Supabase `emails` table, NOT from Gmail. Ingest already persists
every message with its body and its label, so this job needs no mailbox
credentials and does not re-classify anything — it reuses the label the
dashboard shows, which means the report and the UI can never disagree.

Usage:
    python -m backend.app.daily_report                    # yesterday → email
    python -m backend.app.daily_report --dry-run          # save .xlsx locally
    python -m backend.app.daily_report --since 2026-05-01 # backfill from a date

Required env vars:
    SUPABASE_URL, SUPABASE_KEY         (the email store)
    OPENAI_API_KEY                     (field extraction for the spreadsheet)
    EMAIL_ACCOUNT, EMAIL_PASSWORD      (SMTP sender)
    REPORT_RECIPIENT                   (comma-separated recipients)
"""

import argparse
import json
import logging
import os
import re
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from backend.core.config import settings
from backend.core.db import get_db
from backend.core.logging_config import configure_logging, default_log_file

# To a file, not just the console. This runs unattended at 02:30 UTC; without a
# file, a 2am failure leaves nothing behind but whatever the caller happened to
# redirect. (In GitHub Actions LOG_TO_FILE=0 turns this off — the runner's disk
# is discarded and Actions captures stdout anyway.)
configure_logging(log_file=default_log_file())
logger = logging.getLogger(__name__)

settings.require_openai()
openai_client = OpenAI(api_key=settings.openai_api_key)

REPORT_RECIPIENT = settings.report_recipient

REPORTED_LABELS = ("customer_requirement", "quotation_rate_card")
PAGE = 500


# ---------------------------------------------------------------------------
# Email fetching — from the store, not the mailbox
# ---------------------------------------------------------------------------

def _fetch_emails(since: date, before: date | None) -> list[dict]:
    """Every reportable email received in [since, before).

    Only the two labels that appear in the report are selected, so a day of
    routine operational mail costs nothing to skip. Rows whose classification
    never succeeded are counted separately and warned about — they are the
    reason a report might be short.
    """
    start = datetime.combine(since, datetime.min.time(), tzinfo=timezone.utc).isoformat()
    rows: list[dict] = []
    offset = 0
    while True:
        q = (
            get_db().table("emails")
            .select("provider_msg_id, thread_id, sender, subject, body, received_at, classification")
            .in_("classification", list(REPORTED_LABELS))
            .gte("received_at", start)
            .order("received_at")
            .range(offset, offset + PAGE - 1)
        )
        if before:
            end = datetime.combine(before, datetime.min.time(), tzinfo=timezone.utc).isoformat()
            q = q.lt("received_at", end)
        page = q.execute().data or []
        rows.extend(page)
        if len(page) < PAGE:
            break
        offset += PAGE

    _warn_if_unclassified(since, before)

    # One enquiry per thread: a chain of follow-ups is still a single request,
    # and the oldest message is the one that states it.
    seen_threads: set[str] = set()
    emails: list[dict] = []
    for r in rows:  # already oldest-first
        tid = r.get("thread_id") or r["provider_msg_id"]
        if tid in seen_threads:
            continue
        seen_threads.add(tid)
        body = (r.get("body") or "").strip()
        if len(body) < 20:
            continue
        emails.append({
            "id": r["provider_msg_id"],
            "thread_id": tid,
            "from": r.get("sender", ""),
            "subject": r.get("subject", ""),
            "body": body,
            "date": (r.get("received_at") or "")[:10],
            "label": r["classification"],
        })

    logger.info("Reportable emails: %d (%d unique threads)", len(rows), len(emails))
    return emails


def _warn_if_unclassified(since: date, before: date | None) -> None:
    """A pending row is an email the classifier never managed to label. It is
    invisible to this report, so say so rather than quietly under-reporting."""
    try:
        start = datetime.combine(since, datetime.min.time(), tzinfo=timezone.utc).isoformat()
        q = (
            get_db().table("emails").select("id", count="exact")
            .eq("classification_status", "pending")
            .gte("received_at", start).limit(1)
        )
        if before:
            end = datetime.combine(before, datetime.min.time(), tzinfo=timezone.utc).isoformat()
            q = q.lt("received_at", end)
        pending = q.execute().count or 0
        if pending:
            logger.warning(
                "%d email(s) in this window are still unclassified and are NOT in "
                "the report — the retry job will label them later.", pending
            )
    except Exception as e:
        logger.warning("Could not check for unclassified email: %s", e)


# ---------------------------------------------------------------------------
# Field extraction
# ---------------------------------------------------------------------------

_CRQ_PROMPT = """\
Extract structured fields from this logistics enquiry email.
Return ONLY valid JSON with these exact keys (use null if not found):
{{
  "sender_name": "...",
  "company": "...",
  "pol": "...",
  "pod": "...",
  "container_type": "...",
  "container_count": "...",
  "cargo_type": "...",
  "incoterms": "...",
  "remarks": "..."
}}

From: {from_}
Subject: {subject}
Body:
{body}
"""

_RCQ_PROMPT = """\
Extract structured fields from this freight rate card / rate offer email.
Return ONLY valid JSON with these exact keys (use null if not found):
{{
  "carrier_agent": "...",
  "pol": "...",
  "pod": "...",
  "container_type": "...",
  "rate_usd": "...",
  "validity": "...",
  "transit_time": "...",
  "remarks": "..."
}}

From: {from_}
Subject: {subject}
Body:
{body}
"""


def _extract_fields(email: dict, label: str) -> dict:
    """Extract structured fields from a classified email."""
    prompt_tpl = _CRQ_PROMPT if label == "customer_requirement" else _RCQ_PROMPT
    prompt = prompt_tpl.format(
        from_=email["from"],
        subject=email["subject"],
        body=email["body"][:1200],
    )
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=300,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        fields = json.loads(raw)
        fields["subject"] = email["subject"]
        fields["from"] = email["from"]
        fields["date"] = email.get("date", "")
        return fields
    except Exception as e:
        logger.warning("Extract failed for '%s': [%s] %s", email.get("subject", "")[:40], type(e).__name__, e)
        return {
            "subject": email["subject"],
            "from": email["from"],
            "date": email.get("date", ""),
            "error": str(e),
        }


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="D9E1F2")


def _style_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 22


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def _add_rows(ws, rows: list[dict], headers: list[str]) -> None:
    for r_idx, row in enumerate(rows, 2):
        fill = _ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(headers, 1):
            val = row.get(key.lower().replace(" ", "_"), row.get(key, ""))
            cell = ws.cell(row=r_idx, column=c_idx, value=val or "")
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)


_CRQ_HEADERS = [
    "Date", "From", "Subject",
    "Sender Name", "Company",
    "POL", "POD",
    "Container Type", "Container Count",
    "Cargo Type", "Incoterms", "Remarks",
]

_RCQ_HEADERS = [
    "Date", "From", "Subject",
    "Carrier / Agent",
    "POL", "POD",
    "Container Type", "Rate (USD)", "Validity", "Transit Time", "Remarks",
]

_CRQ_KEYS = [
    "date", "from", "subject",
    "sender_name", "company",
    "pol", "pod",
    "container_type", "container_count",
    "cargo_type", "incoterms", "remarks",
]

_RCQ_KEYS = [
    "date", "from", "subject",
    "carrier_agent",
    "pol", "pod",
    "container_type", "rate_usd", "validity", "transit_time", "remarks",
]


def build_excel(crq_rows: list[dict], rcq_rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()

    # Sheet 1: Quote Requests
    ws1 = wb.active
    ws1.title = "Quote Requests"
    ws1.freeze_panes = "A2"
    _style_header(ws1, _CRQ_HEADERS)
    for r_idx, row in enumerate(crq_rows, 2):
        alt = PatternFill("solid", fgColor="D9E1F2") if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(_CRQ_KEYS, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=row.get(key, "") or "")
            if alt:
                cell.fill = alt
            cell.alignment = Alignment(wrap_text=True)
    _auto_width(ws1)

    # Sheet 2: Rate Cards Received
    ws2 = wb.create_sheet("Rate Cards Received")
    ws2.freeze_panes = "A2"
    _style_header(ws2, _RCQ_HEADERS)
    for r_idx, row in enumerate(rcq_rows, 2):
        alt = PatternFill("solid", fgColor="D9E1F2") if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(_RCQ_KEYS, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=row.get(key, "") or "")
            if alt:
                cell.fill = alt
            cell.alignment = Alignment(wrap_text=True)
    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Email sender (with attachment)
# ---------------------------------------------------------------------------

def _send_report(excel_bytes: bytes, today: str, crq_count: int, rcq_count: int) -> None:
    import smtplib
    import ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    account = os.environ["EMAIL_ACCOUNT"]
    password = os.environ["EMAIL_PASSWORD"]
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    # Support comma-separated recipients
    recipients = [r.strip() for r in os.environ["REPORT_RECIPIENT"].split(",") if r.strip()]

    msg = MIMEMultipart()
    msg["From"] = account
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"Daily Logistics Report — {today}"

    body = (
        f"Hi,\n\n"
        f"Please find today's logistics inbox summary attached.\n\n"
        f"  • Quote Requests (from customers/agents): {crq_count}\n"
        f"  • Rate Cards Received (from carriers/agents): {rcq_count}\n\n"
        f"Report date: {today}\n\n"
        f"Regards,\nBhatia Shipping Copilot"
    )
    msg.attach(MIMEText(body, "plain"))

    attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    attachment.set_payload(excel_bytes)
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition",
        f'attachment; filename="logistics_report_{today}.xlsx"',
    )
    msg.attach(attachment)

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
        server.ehlo()
        server.starttls(context=context)
        server.login(account, password)
        server.sendmail(account, recipients, msg.as_string())
    logger.info("Report emailed to %s", ", ".join(recipients))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _parse_args() -> tuple[date, date | None, bool]:
    """Parse CLI args. Returns (since, before, dry_run).

    Default: yesterday only. With --since DATE: everything from that date on.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--since", metavar="YYYY-MM-DD",
                        help="backfill from this date instead of reporting yesterday")
    parser.add_argument("--dry-run", action="store_true",
                        help="write the .xlsx locally instead of emailing it")
    args = parser.parse_args()

    if args.since:
        since = date.fromisoformat(args.since.replace("/", "-"))
        return since, None, args.dry_run

    yesterday = date.today() - timedelta(days=1)
    return yesterday, date.today(), args.dry_run


def main() -> None:
    since, before, dry_run = _parse_args()
    backfill = before is None
    report_date = since.isoformat()

    logger.info("=== Daily Report — %s ===", report_date)
    logger.info("%s: reading stored labels from Supabase, no re-classification",
                f"Backfill since {since}" if backfill else f"Daily mode, {since} only")

    emails = _fetch_emails(since, before)
    if not emails:
        logger.info("No reportable emails in this window. Nothing to send.")
        return

    crq_emails = [e for e in emails if e["label"] == "customer_requirement"]
    rcq_emails = [e for e in emails if e["label"] == "quotation_rate_card"]

    logger.info("Extracting fields from %d quote requests...", len(crq_emails))
    crq_rows = [_extract_fields(em, "customer_requirement") for em in crq_emails]

    logger.info("Extracting fields from %d rate cards...", len(rcq_emails))
    rcq_rows = [_extract_fields(em, "quotation_rate_card") for em in rcq_emails]

    logger.info("Building Excel...")
    excel_bytes = build_excel(crq_rows, rcq_rows)

    filename = f"logistics_report_{report_date}.xlsx"

    if dry_run:
        with open(filename, "wb") as f:
            f.write(excel_bytes)
        logger.info("DRY RUN — saved to %s (not emailed)", filename)
        logger.info("  Quote Requests : %d", len(crq_rows))
        logger.info("  Rate Cards     : %d", len(rcq_rows))
        return

    if not REPORT_RECIPIENT:
        logger.error("REPORT_RECIPIENT not set in .env — cannot email report.")
        with open(filename, "wb") as f:
            f.write(excel_bytes)
        logger.info("Saved locally to %s", filename)
        return

    _send_report(excel_bytes, report_date, len(crq_rows), len(rcq_rows))
    logger.info("Done. Quote Requests: %d | Rate Cards: %d", len(crq_rows), len(rcq_rows))


if __name__ == "__main__":
    main()
